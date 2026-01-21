#!/usr/bin/env python3
"""
Options Pricing & P&L Tool
Fetches options data from Yahoo Finance and creates an Excel file with P&L calculations.

Usage:
    python options_tool.py                    # Interactive mode
    python options_tool.py AAPL               # Fetch AAPL options (nearest expiry)
    python options_tool.py AAPL 2025-01-17    # Fetch AAPL options for specific expiry
"""

import sys
import math
from datetime import datetime
from pathlib import Path

try:
    import yfinance as yf
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing required package: {e.name}")
    print("\nInstall required packages with:")
    print("    pip install yfinance pandas openpyxl")
    sys.exit(1)

# Try to import scipy for normal CDF, fall back to manual implementation
try:
    from scipy.stats import norm
    def normal_cdf(x):
        return norm.cdf(x)
except ImportError:
    def normal_cdf(x):
        """Manual implementation of standard normal CDF using error function approximation."""
        return 0.5 * (1 + math.erf(x / math.sqrt(2)))


# Style constants
DARK_HEADER_FILL = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
BLUE_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
YELLOW_INPUT_FILL = PatternFill(start_color="FFFFC8", end_color="FFFFC8", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Risk-free rate (approximate current rate)
RISK_FREE_RATE = 0.045  # 4.5%


def calculate_delta(spot, strike, time_to_expiry, volatility, option_type="CALL"):
    """
    Calculate option delta using Black-Scholes model.

    Args:
        spot: Current stock price
        strike: Option strike price
        time_to_expiry: Time to expiration in years
        volatility: Implied volatility (as decimal, e.g., 0.30 for 30%)
        option_type: "CALL" or "PUT"

    Returns:
        Delta value
    """
    if time_to_expiry <= 0 or volatility <= 0 or spot <= 0 or strike <= 0:
        return 0.5 if option_type == "CALL" else -0.5

    try:
        d1 = (math.log(spot / strike) + (RISK_FREE_RATE + 0.5 * volatility ** 2) * time_to_expiry) / (volatility * math.sqrt(time_to_expiry))

        if option_type.upper() == "CALL":
            return normal_cdf(d1)
        else:
            return normal_cdf(d1) - 1
    except (ValueError, ZeroDivisionError):
        return 0.5 if option_type == "CALL" else -0.5


def get_options_data(ticker: str, expiry_date: str = None):
    """Fetch options data from Yahoo Finance."""
    stock = yf.Ticker(ticker)

    # Get current stock price
    try:
        current_price = stock.info.get('regularMarketPrice') or stock.info.get('currentPrice')
        if current_price is None:
            hist = stock.history(period="1d")
            current_price = hist['Close'].iloc[-1] if not hist.empty else 0
    except Exception:
        current_price = 0

    # Get available expiration dates
    try:
        expirations = stock.options
    except Exception as e:
        print(f"Error fetching options for {ticker}: {e}")
        return None, None, None, [], current_price

    if not expirations:
        print(f"No options available for {ticker}")
        return None, None, None, [], current_price

    # Select expiration date
    if expiry_date and expiry_date in expirations:
        selected_expiry = expiry_date
    elif expiry_date:
        print(f"Expiry {expiry_date} not available. Available dates:")
        for exp in expirations[:10]:
            print(f"  {exp}")
        selected_expiry = expirations[0]
        print(f"Using nearest: {selected_expiry}")
    else:
        selected_expiry = expirations[0]

    # Get options chain
    try:
        chain = stock.option_chain(selected_expiry)
        calls_df = chain.calls
        puts_df = chain.puts
    except Exception as e:
        print(f"Error fetching options chain: {e}")
        return None, None, None, expirations, current_price

    return calls_df, puts_df, selected_expiry, expirations, current_price


def create_options_sheet(ws, options_df: pd.DataFrame, option_type: str,
                         expiry: str, current_price: float, ticker: str):
    """Create a formatted sheet with options data and P&L formulas."""

    # Calculate time to expiry in years
    expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
    today = datetime.now()
    days_to_expiry = (expiry_date - today).days
    time_to_expiry = max(days_to_expiry / 365.0, 0.001)  # Avoid zero

    # === INPUT AREA (Rows 1-2) ===
    headers_row1 = [
        (2, "Ticker"),
        (3, "Expiry"),
        (4, "Days to Exp"),
        (6, "Current Price"),
        (8, "Stock @ Expiry"),
    ]

    for col, text in headers_row1:
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = DARK_HEADER_FILL if col <= 6 else BLUE_HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER

    # Input values (Row 2)
    ws.cell(row=2, column=2, value=ticker.upper()).border = THIN_BORDER
    ws.cell(row=2, column=3, value=expiry).border = THIN_BORDER
    ws.cell(row=2, column=4, value=days_to_expiry).border = THIN_BORDER

    price_cell = ws.cell(row=2, column=6, value=current_price)
    price_cell.number_format = '$#,##0.00'
    price_cell.border = THIN_BORDER

    # Stock @ Expiry input cell
    expiry_cell = ws.cell(row=2, column=8, value=current_price)
    expiry_cell.number_format = '$#,##0.00'
    expiry_cell.fill = YELLOW_INPUT_FILL
    expiry_cell.border = THIN_BORDER

    # === P&L SUMMARY (Row 3) ===
    ws.cell(row=3, column=12, value="Total P&L:").font = Font(bold=True)

    # === COLUMN HEADERS (Row 5) ===
    data_headers = [
        ("Strike", "$#,##0.00"),
        ("Bid", "$#,##0.00"),
        ("Ask", "$#,##0.00"),
        ("Last", "$#,##0.00"),
        ("Mid", "$#,##0.00"),
        ("Volume", "#,##0"),
        ("Open Int", "#,##0"),
        ("Impl Vol", "0.00%"),
        ("Delta", "0.00"),
        ("Position", "0"),
        ("Entry Price", "$#,##0.00"),
        ("Value@Exp", "$#,##0.00"),
        ("P&L", "$#,##0.00"),
    ]

    for col, (header, _) in enumerate(data_headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        if col <= 9:
            cell.fill = DARK_HEADER_FILL
        else:
            cell.fill = BLUE_HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER

    # === OPTIONS DATA (Row 6+) ===
    data_start_row = 6

    for i, (_, row_data) in enumerate(options_df.iterrows()):
        row = data_start_row + i

        strike = row_data.get('strike', 0)
        bid = row_data.get('bid', 0) or 0
        ask = row_data.get('ask', 0) or 0
        last = row_data.get('lastPrice', 0) or 0
        mid = (bid + ask) / 2 if bid > 0 and ask > 0 else last
        volume = row_data.get('volume', 0) or 0
        oi = row_data.get('openInterest', 0) or 0
        iv = row_data.get('impliedVolatility', 0) or 0

        # Market data columns
        ws.cell(row=row, column=1, value=strike).number_format = '$#,##0.00'
        ws.cell(row=row, column=2, value=bid).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=ask).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=last).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=mid).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=int(volume) if pd.notna(volume) else 0).number_format = '#,##0'
        ws.cell(row=row, column=7, value=int(oi) if pd.notna(oi) else 0).number_format = '#,##0'
        ws.cell(row=row, column=8, value=iv if pd.notna(iv) else 0).number_format = '0.00%'

        # Calculate delta using Black-Scholes
        if pd.notna(iv) and iv > 0:
            delta = calculate_delta(current_price, strike, time_to_expiry, iv, option_type)
        else:
            delta = 0
        ws.cell(row=row, column=9, value=delta).number_format = '0.00'

        # Position (user input, default 0)
        pos_cell = ws.cell(row=row, column=10, value=0)
        pos_cell.number_format = '0'
        pos_cell.fill = YELLOW_INPUT_FILL

        # Entry Price (user input, default to mid)
        entry_cell = ws.cell(row=row, column=11, value=mid)
        entry_cell.number_format = '$#,##0.00'
        entry_cell.fill = YELLOW_INPUT_FILL

        # Value at Expiry formula
        if option_type.upper() == "CALL":
            value_formula = f"=MAX($H$2-A{row},0)"
        else:
            value_formula = f"=MAX(A{row}-$H$2,0)"
        ws.cell(row=row, column=12, value=value_formula).number_format = '$#,##0.00'

        # P&L formula: =(L6-K6)*J6*100
        pnl_formula = f"=(L{row}-K{row})*J{row}*100"
        ws.cell(row=row, column=13, value=pnl_formula).number_format = '$#,##0.00'

    last_data_row = data_start_row + len(options_df) - 1

    # Total P&L formula
    total_pnl_cell = ws.cell(row=3, column=13, value=f"=SUM(M{data_start_row}:M{last_data_row})")
    total_pnl_cell.number_format = '$#,##0.00'
    total_pnl_cell.font = Font(bold=True)

    # Add conditional formatting for P&L column
    green_font = Font(color="008000")
    red_font = Font(color="C00000")

    pnl_range = f"M{data_start_row}:M{last_data_row}"
    ws.conditional_formatting.add(pnl_range,
        CellIsRule(operator='greaterThan', formula=['0'], font=green_font))
    ws.conditional_formatting.add(pnl_range,
        CellIsRule(operator='lessThan', formula=['0'], font=red_font))

    ws.conditional_formatting.add("M3",
        CellIsRule(operator='greaterThan', formula=['0'], font=green_font))
    ws.conditional_formatting.add("M3",
        CellIsRule(operator='lessThan', formula=['0'], font=red_font))

    # Auto-fit column widths
    column_widths = [10, 8, 8, 8, 8, 10, 10, 10, 8, 10, 12, 10, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze panes
    ws.freeze_panes = 'A6'


def create_excel_file(ticker: str, calls_df: pd.DataFrame, puts_df: pd.DataFrame,
                      expiry: str, current_price: float, expirations: list, output_path: str):
    """Create formatted Excel file with Calls and Puts on separate tabs."""

    wb = Workbook()

    # Create Calls sheet (rename default sheet)
    ws_calls = wb.active
    ws_calls.title = "Calls"
    create_options_sheet(ws_calls, calls_df, "CALL", expiry, current_price, ticker)

    # Create Puts sheet
    ws_puts = wb.create_sheet("Puts")
    create_options_sheet(ws_puts, puts_df, "PUT", expiry, current_price, ticker)

    # Create Expirations reference sheet
    ws_exp = wb.create_sheet("Expirations")
    ws_exp.cell(row=1, column=1, value="Available Expiration Dates").font = Font(bold=True)
    ws_exp.cell(row=1, column=2, value=f"for {ticker.upper()}").font = Font(bold=True)
    ws_exp.cell(row=2, column=1, value="(Re-run script with desired date)").font = Font(italic=True)

    for i, exp_date in enumerate(expirations, start=4):
        cell = ws_exp.cell(row=i, column=1, value=exp_date)
        if exp_date == expiry:
            cell.font = Font(bold=True, color="0070C0")
            ws_exp.cell(row=i, column=2, value="← Current").font = Font(color="0070C0")

    ws_exp.column_dimensions['A'].width = 15
    ws_exp.column_dimensions['B'].width = 12

    # Save
    wb.save(output_path)
    return output_path


def list_expirations(ticker: str):
    """List available expiration dates for a ticker."""
    stock = yf.Ticker(ticker)
    try:
        expirations = stock.options
        print(f"\nAvailable expiration dates for {ticker.upper()}:")
        for exp in expirations:
            print(f"  {exp}")
        return expirations
    except Exception as e:
        print(f"Error: {e}")
        return []


def main():
    print("=" * 50)
    print("OPTIONS PRICING & P&L TOOL")
    print("=" * 50)

    # Parse command line arguments
    if len(sys.argv) >= 2:
        ticker = sys.argv[1].upper()
        expiry = sys.argv[2] if len(sys.argv) >= 3 else None
    else:
        # Interactive mode
        ticker = input("\nEnter ticker symbol (e.g., AAPL): ").strip().upper()
        if not ticker:
            ticker = "AAPL"
        expiry = None

    # Always show expirations and let user pick if not specified
    print(f"\nFetching available expirations for {ticker}...")
    stock = yf.Ticker(ticker)

    try:
        expirations = stock.options
    except Exception as e:
        print(f"Error: {e}")
        return

    if not expirations:
        print(f"No options available for {ticker}")
        return

    # Show numbered list of expirations
    print(f"\nAvailable expiration dates for {ticker}:")
    print("-" * 35)
    for i, exp in enumerate(expirations, start=1):
        print(f"  {i:2}. {exp}")
    print("-" * 35)

    # Let user pick if expiry not specified or invalid
    if expiry and expiry in expirations:
        selected_expiry = expiry
        print(f"\nUsing specified expiry: {selected_expiry}")
    else:
        if expiry:
            print(f"\n'{expiry}' not found in available dates.")

        choice = input(f"\nEnter number (1-{len(expirations)}) or date (YYYY-MM-DD) [default: 1]: ").strip()

        if not choice:
            selected_expiry = expirations[0]
        elif choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(expirations):
                selected_expiry = expirations[idx]
            else:
                print("Invalid choice, using first expiration.")
                selected_expiry = expirations[0]
        elif choice in expirations:
            selected_expiry = choice
        else:
            print("Invalid choice, using first expiration.")
            selected_expiry = expirations[0]

    print(f"\nFetching {ticker} options for {selected_expiry}...")

    # Fetch options data
    calls_df, puts_df, _, _, current_price = get_options_data(ticker, selected_expiry)

    if calls_df is None or calls_df.empty:
        print("No options data retrieved.")
        return

    print(f"Current stock price: ${current_price:.2f}")
    print(f"Found {len(calls_df)} calls and {len(puts_df)} puts")
    print(f"Deltas calculated using Black-Scholes (r={RISK_FREE_RATE*100:.1f}%)")

    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"{ticker}_OPTIONS_{selected_expiry}_{timestamp}.xlsx"
    output_path = Path.cwd() / output_file

    # Create Excel file
    create_excel_file(
        ticker=ticker,
        calls_df=calls_df,
        puts_df=puts_df,
        expiry=selected_expiry,
        current_price=current_price,
        expirations=expirations,
        output_path=str(output_path)
    )

    print(f"\n{'=' * 50}")
    print(f"Excel file created: {output_path}")
    print(f"{'=' * 50}")
    print("\nTabs: 'Calls', 'Puts', and 'Expirations' (reference list)")
    print("\nHOW TO USE:")
    print("1. Open the Excel file")
    print("2. Enter positions in the yellow 'Position' column (+long, -short)")
    print("3. Adjust 'Entry Price' if different from mid")
    print("4. Change 'Stock @ Expiry' (cell H2) to see P&L at different prices")
    print("5. Total P&L shows in cell M3 on each tab")
    print(f"\nTo get a different expiration, run: python options_tool.py {ticker}")


if __name__ == "__main__":
    main()
