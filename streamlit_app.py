"""
Options Pricing & P&L Tool - Streamlit Web App
Interactive P&L calculator with position entry.
"""

import streamlit as st
import pandas as pd
import yfinance as yf
import math
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# Try scipy, fall back to math.erf
try:
    from scipy.stats import norm
    def normal_cdf(x):
        return norm.cdf(x)
except ImportError:
    def normal_cdf(x):
        return 0.5 * (1 + math.erf(x / math.sqrt(2)))

# Constants
RISK_FREE_RATE = 0.045

st.set_page_config(page_title="Options Pricing Tool", page_icon="📈", layout="wide")


# --- Authentication ---
def check_password():
    """Simple password protection."""

    # Check if already authenticated
    if st.session_state.get("authenticated", False):
        return True

    # Get password from secrets or use default for local dev
    try:
        correct_password = st.secrets["password"]
    except (KeyError, FileNotFoundError):
        # No password set - allow access (for local development)
        return True

    st.title("📈 Options Pricing Tool")
    st.markdown("### Login Required")

    password = st.text_input("Password", type="password", key="password_input")

    if st.button("Login", type="primary"):
        if password == correct_password:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password")

    return False


# Check authentication before showing app
if not check_password():
    st.stop()

st.title("📈 Options Pricing & P&L Tool")


def calculate_delta(spot, strike, time_to_expiry, volatility, option_type="CALL"):
    if time_to_expiry <= 0 or volatility <= 0 or spot <= 0 or strike <= 0:
        return 0.5 if option_type == "CALL" else -0.5
    try:
        d1 = (math.log(spot / strike) + (RISK_FREE_RATE + 0.5 * volatility ** 2) * time_to_expiry) / (volatility * math.sqrt(time_to_expiry))
        return normal_cdf(d1) if option_type == "CALL" else normal_cdf(d1) - 1
    except:
        return 0.5 if option_type == "CALL" else -0.5


def black_scholes_price(spot, strike, time_to_expiry, volatility, option_type="CALL"):
    """Calculate Black-Scholes option price."""
    if time_to_expiry <= 0:
        # At expiry, return intrinsic value
        if option_type == "CALL":
            return max(spot - strike, 0)
        else:
            return max(strike - spot, 0)

    if volatility <= 0 or spot <= 0 or strike <= 0:
        return 0

    try:
        d1 = (math.log(spot / strike) + (RISK_FREE_RATE + 0.5 * volatility ** 2) * time_to_expiry) / (volatility * math.sqrt(time_to_expiry))
        d2 = d1 - volatility * math.sqrt(time_to_expiry)

        if option_type == "CALL":
            price = spot * normal_cdf(d1) - strike * math.exp(-RISK_FREE_RATE * time_to_expiry) * normal_cdf(d2)
        else:
            price = strike * math.exp(-RISK_FREE_RATE * time_to_expiry) * normal_cdf(-d2) - spot * normal_cdf(-d1)

        return max(price, 0)
    except:
        return 0


@st.cache_data(ttl=300)
def get_expirations_and_price(ticker):
    """Fetch stock price and available expirations."""
    stock = yf.Ticker(ticker)
    try:
        expirations = list(stock.options)
    except:
        return [], 0

    try:
        price = stock.info.get('regularMarketPrice') or stock.info.get('currentPrice') or 0
        if not price:
            hist = stock.history(period="1d")
            price = hist['Close'].iloc[-1] if not hist.empty else 0
    except:
        price = 0

    return expirations, price


@st.cache_data(ttl=300)
def get_options_chain(ticker, expiry):
    """Fetch options chain for given expiry."""
    stock = yf.Ticker(ticker)
    try:
        chain = stock.option_chain(expiry)
        return chain.calls.to_dict('records'), chain.puts.to_dict('records')
    except:
        return None, None


def process_options_df(records, current_price, time_to_expiry, option_type):
    """Process options data and calculate delta."""
    result = []
    for row in records:
        strike = row.get('strike', 0)
        bid = row.get('bid', 0) or 0
        ask = row.get('ask', 0) or 0
        last = row.get('lastPrice', 0) or 0
        mid = (bid + ask) / 2 if bid > 0 and ask > 0 else last
        volume = row.get('volume', 0) or 0
        oi = row.get('openInterest', 0) or 0
        iv = row.get('impliedVolatility', 0) or 0

        delta = calculate_delta(current_price, strike, time_to_expiry, iv, option_type) if iv > 0 else 0

        result.append({
            'Strike': strike,
            'Bid': bid,
            'Ask': ask,
            'Last': last,
            'Mid': mid,
            'Volume': int(volume) if pd.notna(volume) else 0,
            'OpenInt': int(oi) if pd.notna(oi) else 0,
            'IV': iv,
            'Delta': round(delta, 2),
            'Position': 0,
            'Entry': round(mid, 2),
        })

    return pd.DataFrame(result)


def calculate_pnl(df, stock_scenario, option_type, days_to_scenario=0, iv_adjustment=0.0):
    """Calculate P&L for each position.

    Args:
        df: DataFrame with options data
        stock_scenario: Stock price at scenario date
        option_type: "CALL" or "PUT"
        days_to_scenario: Days until scenario (0 = at expiry, >0 = before expiry)
        iv_adjustment: Percentage adjustment to IV (e.g., 0.1 = +10% IV)
    """
    df = df.copy()

    if days_to_scenario <= 0:
        # At expiry - use intrinsic value
        if option_type == "CALL":
            df['ScenValue'] = (stock_scenario - df['Strike']).clip(lower=0)
        else:
            df['ScenValue'] = (df['Strike'] - stock_scenario).clip(lower=0)
    else:
        # Before expiry - use Black-Scholes
        time_to_expiry = days_to_scenario / 365.0
        df['ScenValue'] = df.apply(
            lambda r: black_scholes_price(
                stock_scenario,
                r['Strike'],
                time_to_expiry,
                r['IV'] * (1 + iv_adjustment) if r['IV'] > 0 else 0.3,  # Default 30% IV if missing
                option_type
            ),
            axis=1
        )

    df['PremPaid'] = df.apply(lambda r: r['Entry'] * r['Position'] * 100 if r['Position'] > 0 else 0, axis=1)
    df['PremRcvd'] = df.apply(lambda r: r['Entry'] * -r['Position'] * 100 if r['Position'] < 0 else 0, axis=1)
    df['Payout'] = df['ScenValue'] * df['Position'] * 100
    df['P&L'] = df['Payout'] - df['PremPaid'] + df['PremRcvd']

    return df


def create_excel_download(ticker, calls_df, puts_df, expiry, current_price, expirations):
    """Create Excel file for download."""
    DARK_HEADER_FILL = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
    BLUE_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    YELLOW_INPUT_FILL = PatternFill(start_color="FFFFC8", end_color="FFFFC8", fill_type="solid")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def create_sheet(ws, options_df, option_type):
        expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
        days_to_expiry = (expiry_date - datetime.now()).days

        for col, text in [(2, "Ticker"), (3, "Expiry"), (4, "Days"), (6, "Current Price"), (8, "Stock @ Expiry")]:
            cell = ws.cell(row=1, column=col, value=text)
            cell.fill = DARK_HEADER_FILL if col <= 6 else BLUE_HEADER_FILL
            cell.font = WHITE_FONT
            cell.border = THIN_BORDER

        ws.cell(row=2, column=2, value=ticker).border = THIN_BORDER
        ws.cell(row=2, column=3, value=expiry).border = THIN_BORDER
        ws.cell(row=2, column=4, value=days_to_expiry).border = THIN_BORDER
        ws.cell(row=2, column=6, value=current_price).number_format = '$#,##0.00'
        ws.cell(row=2, column=6).border = THIN_BORDER

        exp_cell = ws.cell(row=2, column=8, value=current_price)
        exp_cell.number_format = '$#,##0.00'
        exp_cell.fill = YELLOW_INPUT_FILL
        exp_cell.border = THIN_BORDER

        ws.cell(row=1, column=10, value="Stock Shares").fill = BLUE_HEADER_FILL
        ws.cell(row=1, column=10).font = WHITE_FONT
        ws.cell(row=1, column=10).border = THIN_BORDER
        ws.cell(row=1, column=11, value="Stock Entry").fill = BLUE_HEADER_FILL
        ws.cell(row=1, column=11).font = WHITE_FONT
        ws.cell(row=1, column=11).border = THIN_BORDER
        ws.cell(row=1, column=12, value="Stock P&L").fill = BLUE_HEADER_FILL
        ws.cell(row=1, column=12).font = WHITE_FONT
        ws.cell(row=1, column=12).border = THIN_BORDER

        ws.cell(row=2, column=10, value=0).fill = YELLOW_INPUT_FILL
        ws.cell(row=2, column=10).border = THIN_BORDER
        ws.cell(row=2, column=11, value=current_price).number_format = '$#,##0.00'
        ws.cell(row=2, column=11).fill = YELLOW_INPUT_FILL
        ws.cell(row=2, column=11).border = THIN_BORDER
        ws.cell(row=2, column=12, value="=(H2-K2)*J2").number_format = '$#,##0.00'
        ws.cell(row=2, column=12).border = THIN_BORDER
        ws.cell(row=2, column=12).font = Font(bold=True)

        ws.cell(row=3, column=1, value="SUMMARY").font = Font(bold=True, size=11)
        ws.cell(row=3, column=2, value="Premiums Paid:").font = Font(bold=True)
        ws.cell(row=3, column=4, value="Premiums Rcvd:").font = Font(bold=True)
        ws.cell(row=3, column=6, value="Options Payout:").font = Font(bold=True)
        ws.cell(row=3, column=8, value="Options P&L:").font = Font(bold=True)
        ws.cell(row=3, column=10, value="Stock P&L:").font = Font(bold=True)
        ws.cell(row=3, column=11, value="=L2").number_format = '$#,##0.00'
        ws.cell(row=3, column=11).font = Font(bold=True)
        ws.cell(row=3, column=12, value="TOTAL P&L:").font = Font(bold=True, color="0070C0")

        headers = ["Strike", "Bid", "Ask", "Last", "Mid", "Volume", "Open Int", "Impl Vol", "Delta",
                   "Position", "Entry", "Prem Paid", "Prem Rcvd", "Val@Exp", "Payout", "P&L"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.fill = DARK_HEADER_FILL if col <= 9 else BLUE_HEADER_FILL
            cell.font = WHITE_FONT
            cell.border = THIN_BORDER

        row = 6
        for _, r in options_df.iterrows():
            ws.cell(row=row, column=1, value=r['Strike']).number_format = '$#,##0.00'
            ws.cell(row=row, column=2, value=r['Bid']).number_format = '$#,##0.00'
            ws.cell(row=row, column=3, value=r['Ask']).number_format = '$#,##0.00'
            ws.cell(row=row, column=4, value=r['Last']).number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value=r['Mid']).number_format = '$#,##0.00'
            ws.cell(row=row, column=6, value=r['Volume']).number_format = '#,##0'
            ws.cell(row=row, column=7, value=r['OpenInt']).number_format = '#,##0'
            ws.cell(row=row, column=8, value=r['IV']).number_format = '0.00%'
            ws.cell(row=row, column=9, value=r['Delta']).number_format = '0.00'

            ws.cell(row=row, column=10, value=r.get('Position', 0)).fill = YELLOW_INPUT_FILL
            ws.cell(row=row, column=11, value=r.get('Entry', r['Mid'])).number_format = '$#,##0.00'
            ws.cell(row=row, column=11).fill = YELLOW_INPUT_FILL

            ws.cell(row=row, column=12, value=f"=IF(J{row}>0,K{row}*J{row}*100,0)").number_format = '$#,##0.00'
            ws.cell(row=row, column=13, value=f"=IF(J{row}<0,K{row}*-J{row}*100,0)").number_format = '$#,##0.00'

            if option_type == "CALL":
                ws.cell(row=row, column=14, value=f"=MAX($H$2-A{row},0)").number_format = '$#,##0.00'
            else:
                ws.cell(row=row, column=14, value=f"=MAX(A{row}-$H$2,0)").number_format = '$#,##0.00'

            ws.cell(row=row, column=15, value=f"=N{row}*J{row}*100").number_format = '$#,##0.00'
            ws.cell(row=row, column=16, value=f"=O{row}-L{row}+M{row}").number_format = '$#,##0.00'

            row += 1

        last_row = row - 1

        ws.cell(row=3, column=3, value=f"=SUM(L6:L{last_row})").number_format = '$#,##0.00'
        ws.cell(row=3, column=3).font = Font(bold=True)
        ws.cell(row=3, column=5, value=f"=SUM(M6:M{last_row})").number_format = '$#,##0.00'
        ws.cell(row=3, column=5).font = Font(bold=True)
        ws.cell(row=3, column=7, value=f"=SUM(O6:O{last_row})").number_format = '$#,##0.00'
        ws.cell(row=3, column=7).font = Font(bold=True)
        ws.cell(row=3, column=9, value=f"=SUM(P6:P{last_row})").number_format = '$#,##0.00'
        ws.cell(row=3, column=9).font = Font(bold=True)
        ws.cell(row=3, column=13, value="=I3+K3").number_format = '$#,##0.00'
        ws.cell(row=3, column=13).font = Font(bold=True, size=12, color="0070C0")

        green_font = Font(color="008000")
        red_font = Font(color="C00000")
        ws.conditional_formatting.add(f"P6:P{last_row}", CellIsRule(operator='greaterThan', formula=['0'], font=green_font))
        ws.conditional_formatting.add(f"P6:P{last_row}", CellIsRule(operator='lessThan', formula=['0'], font=red_font))

        for i, w in enumerate([10, 8, 8, 8, 8, 9, 9, 9, 7, 8, 8, 10, 10, 9, 10, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A6'

    wb = Workbook()
    ws_calls = wb.active
    ws_calls.title = "Calls"
    create_sheet(ws_calls, calls_df, "CALL")

    ws_puts = wb.create_sheet("Puts")
    create_sheet(ws_puts, puts_df, "PUT")

    ws_exp = wb.create_sheet("Expirations")
    ws_exp.cell(row=1, column=1, value="Available Expirations").font = Font(bold=True)
    for i, exp in enumerate(expirations, 3):
        cell = ws_exp.cell(row=i, column=1, value=exp)
        if exp == expiry:
            cell.font = Font(bold=True, color="0070C0")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- Initialize Session State ---
if 'calls_positions' not in st.session_state:
    st.session_state['calls_positions'] = {}
if 'puts_positions' not in st.session_state:
    st.session_state['puts_positions'] = {}
if 'stock_shares' not in st.session_state:
    st.session_state['stock_shares'] = 0
if 'stock_entry' not in st.session_state:
    st.session_state['stock_entry'] = 0.0


# --- Sidebar ---
with st.sidebar:
    st.header("⚙️ Settings")

    ticker = st.text_input("Ticker Symbol", value="CCJ").upper()

    if st.button("Load Expirations", type="primary"):
        with st.spinner(f"Loading {ticker}..."):
            expirations, price = get_expirations_and_price(ticker)
            if expirations:
                st.session_state['expirations'] = expirations
                st.session_state['current_price'] = price
                st.session_state['ticker'] = ticker
                st.session_state['stock_entry'] = price
                # Reset positions when loading new ticker
                st.session_state['calls_positions'] = {}
                st.session_state['puts_positions'] = {}
                st.session_state['stock_shares'] = 0
                st.success(f"Found {len(expirations)} expirations")
            else:
                st.error("No options found for this ticker")

    if 'expirations' in st.session_state and st.session_state['expirations']:
        expiry = st.selectbox("Expiration Date", st.session_state['expirations'])
        st.session_state['selected_expiry'] = expiry

        st.metric("Current Price", f"${st.session_state.get('current_price', 0):.2f}")

        st.divider()
        st.subheader("📊 Scenario Analysis")

        # Calculate days to expiry for this option
        expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
        actual_days_to_expiry = max((expiry_date - datetime.now()).days, 0)

        days_to_scenario = st.slider(
            "Days to Scenario",
            min_value=0,
            max_value=max(actual_days_to_expiry, 1),
            value=0,
            help="0 = At expiry (intrinsic value), >0 = Before expiry (Black-Scholes theoretical value)"
        )
        st.session_state['days_to_scenario'] = days_to_scenario

        if days_to_scenario > 0:
            st.caption(f"📅 Modeling {days_to_scenario} days before expiry")
        else:
            st.caption("📅 At expiration (intrinsic value)")

        stock_scenario = st.number_input(
            "Stock Price @ Scenario",
            value=st.session_state.get('current_price', 100.0),
            step=1.0,
            format="%.2f"
        )
        st.session_state['stock_scenario'] = stock_scenario

        iv_adjustment = st.slider(
            "IV Adjustment",
            min_value=-0.5,
            max_value=1.0,
            value=0.0,
            step=0.05,
            format="%.0f%%",
            help="Adjust implied volatility for scenario (+10% = IV increases by 10%)"
        )
        st.session_state['iv_adjustment'] = iv_adjustment
        if iv_adjustment != 0:
            st.caption(f"📈 IV {'increased' if iv_adjustment > 0 else 'decreased'} by {abs(iv_adjustment)*100:.0f}%")

        st.divider()
        st.subheader("📈 Stock Position")

        stock_shares = st.number_input(
            "Shares (+long / -short)",
            value=st.session_state.get('stock_shares', 0),
            step=100,
            format="%d"
        )
        st.session_state['stock_shares'] = stock_shares

        stock_entry = st.number_input(
            "Stock Entry Price",
            value=st.session_state.get('stock_entry', st.session_state.get('current_price', 100.0)),
            step=0.5,
            format="%.2f"
        )
        st.session_state['stock_entry'] = stock_entry


# --- Main Content ---
if 'expirations' in st.session_state and st.session_state['expirations']:
    ticker = st.session_state.get('ticker', '')
    expiry = st.session_state.get('selected_expiry', '')
    current_price = st.session_state.get('current_price', 0)
    stock_scenario = st.session_state.get('stock_scenario', current_price)
    days_to_scenario = st.session_state.get('days_to_scenario', 0)
    iv_adjustment = st.session_state.get('iv_adjustment', 0.0)

    expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
    days_to_expiry = (expiry_date - datetime.now()).days
    time_to_expiry = max(days_to_expiry / 365.0, 0.001)

    # Display header with scenario info
    if days_to_scenario > 0:
        st.subheader(f"{ticker} Options - Expires {expiry} ({days_to_expiry} days)")
        st.info(f"🔮 **Scenario Analysis**: {days_to_scenario} days before expiry | Stock @ ${stock_scenario:.2f} | IV adjustment: {iv_adjustment*100:+.0f}%")
    else:
        st.subheader(f"{ticker} Options - Expires {expiry} ({days_to_expiry} days)")

    # Fetch options chain
    with st.spinner("Loading options chain..."):
        calls_raw, puts_raw = get_options_chain(ticker, expiry)

    if calls_raw is not None:
        calls_df = process_options_df(calls_raw, current_price, time_to_expiry, "CALL")
        puts_df = process_options_df(puts_raw, current_price, time_to_expiry, "PUT")

        # Apply saved positions
        for strike, pos in st.session_state.get('calls_positions', {}).items():
            mask = calls_df['Strike'] == strike
            if mask.any():
                calls_df.loc[mask, 'Position'] = pos['position']
                calls_df.loc[mask, 'Entry'] = pos['entry']

        for strike, pos in st.session_state.get('puts_positions', {}).items():
            mask = puts_df['Strike'] == strike
            if mask.any():
                puts_df.loc[mask, 'Position'] = pos['position']
                puts_df.loc[mask, 'Entry'] = pos['entry']

        # Calculate P&L
        calls_pnl = calculate_pnl(calls_df, stock_scenario, "CALL", days_to_scenario, iv_adjustment)
        puts_pnl = calculate_pnl(puts_df, stock_scenario, "PUT", days_to_scenario, iv_adjustment)

        # Stock P&L
        stock_shares = st.session_state.get('stock_shares', 0)
        stock_entry = st.session_state.get('stock_entry', current_price)
        stock_pnl = (stock_scenario - stock_entry) * stock_shares

        # Summary metrics
        calls_with_pos = calls_pnl[calls_pnl['Position'] != 0]
        puts_with_pos = puts_pnl[puts_pnl['Position'] != 0]

        total_prem_paid = calls_with_pos['PremPaid'].sum() + puts_with_pos['PremPaid'].sum()
        total_prem_rcvd = calls_with_pos['PremRcvd'].sum() + puts_with_pos['PremRcvd'].sum()
        total_payout = calls_with_pos['Payout'].sum() + puts_with_pos['Payout'].sum()
        options_pnl = calls_with_pos['P&L'].sum() + puts_with_pos['P&L'].sum()
        total_pnl = options_pnl + stock_pnl

        # Display summary
        if days_to_scenario > 0:
            st.markdown(f"### 💰 P&L Summary (Scenario: {days_to_scenario} days to expiry)")
        else:
            st.markdown("### 💰 P&L Summary (At Expiration)")
        col1, col2, col3, col4, col5 = st.columns(5)
        col1.metric("Premiums Paid", f"${total_prem_paid:,.0f}")
        col2.metric("Premiums Rcvd", f"${total_prem_rcvd:,.0f}")
        col3.metric("Options Payout", f"${total_payout:,.0f}")
        col4.metric("Stock P&L", f"${stock_pnl:,.0f}", delta=f"{stock_shares} shares")
        col5.metric("TOTAL P&L", f"${total_pnl:,.0f}", delta="profit" if total_pnl > 0 else "loss" if total_pnl < 0 else None)

        st.divider()

        # Tabs for Calls and Puts
        tab1, tab2 = st.tabs(["📈 Calls", "📉 Puts"])

        with tab1:
            st.markdown("**Enter positions below** (+ = long, - = short)")

            # Filter to show only relevant strikes (near the money)
            atm_idx = (calls_df['Strike'] - current_price).abs().idxmin()
            start_idx = max(0, atm_idx - 15)
            end_idx = min(len(calls_df), atm_idx + 16)
            display_calls = calls_pnl.iloc[start_idx:end_idx].copy()

            # Dynamic column label based on scenario type
            value_label = "BS Value" if days_to_scenario > 0 else "Val@Exp"

            edited_calls = st.data_editor(
                display_calls[['Strike', 'Bid', 'Ask', 'Mid', 'Delta', 'IV', 'Position', 'Entry', 'PremPaid', 'PremRcvd', 'ScenValue', 'Payout', 'P&L']],
                column_config={
                    "Strike": st.column_config.NumberColumn("Strike", format="$%.2f"),
                    "Bid": st.column_config.NumberColumn("Bid", format="$%.2f"),
                    "Ask": st.column_config.NumberColumn("Ask", format="$%.2f"),
                    "Mid": st.column_config.NumberColumn("Mid", format="$%.2f"),
                    "Delta": st.column_config.NumberColumn("Delta", format="%.2f"),
                    "IV": st.column_config.NumberColumn("IV", format="%.1%%"),
                    "Position": st.column_config.NumberColumn("Position", format="%d", min_value=-100, max_value=100),
                    "Entry": st.column_config.NumberColumn("Entry", format="$%.2f"),
                    "PremPaid": st.column_config.NumberColumn("Prem Paid", format="$%.0f", disabled=True),
                    "PremRcvd": st.column_config.NumberColumn("Prem Rcvd", format="$%.0f", disabled=True),
                    "ScenValue": st.column_config.NumberColumn(value_label, format="$%.2f", disabled=True),
                    "Payout": st.column_config.NumberColumn("Payout", format="$%.0f", disabled=True),
                    "P&L": st.column_config.NumberColumn("P&L", format="$%.0f", disabled=True),
                },
                hide_index=True,
                use_container_width=True,
                key="calls_editor"
            )

            # Save positions and detect changes
            calls_changed = False
            for idx, row in edited_calls.iterrows():
                strike = row['Strike']
                old_pos = st.session_state['calls_positions'].get(strike, {})
                if row['Position'] != 0:
                    new_pos = {'position': row['Position'], 'entry': row['Entry']}
                    if old_pos != new_pos:
                        calls_changed = True
                    st.session_state['calls_positions'][strike] = new_pos
                elif strike in st.session_state['calls_positions']:
                    calls_changed = True
                    del st.session_state['calls_positions'][strike]

            if calls_changed:
                st.rerun()

        with tab2:
            st.markdown("**Enter positions below** (+ = long, - = short)")

            atm_idx = (puts_df['Strike'] - current_price).abs().idxmin()
            start_idx = max(0, atm_idx - 15)
            end_idx = min(len(puts_df), atm_idx + 16)
            display_puts = puts_pnl.iloc[start_idx:end_idx].copy()

            edited_puts = st.data_editor(
                display_puts[['Strike', 'Bid', 'Ask', 'Mid', 'Delta', 'IV', 'Position', 'Entry', 'PremPaid', 'PremRcvd', 'ScenValue', 'Payout', 'P&L']],
                column_config={
                    "Strike": st.column_config.NumberColumn("Strike", format="$%.2f"),
                    "Bid": st.column_config.NumberColumn("Bid", format="$%.2f"),
                    "Ask": st.column_config.NumberColumn("Ask", format="$%.2f"),
                    "Mid": st.column_config.NumberColumn("Mid", format="$%.2f"),
                    "Delta": st.column_config.NumberColumn("Delta", format="%.2f"),
                    "IV": st.column_config.NumberColumn("IV", format="%.1%%"),
                    "Position": st.column_config.NumberColumn("Position", format="%d", min_value=-100, max_value=100),
                    "Entry": st.column_config.NumberColumn("Entry", format="$%.2f"),
                    "PremPaid": st.column_config.NumberColumn("Prem Paid", format="$%.0f", disabled=True),
                    "PremRcvd": st.column_config.NumberColumn("Prem Rcvd", format="$%.0f", disabled=True),
                    "ScenValue": st.column_config.NumberColumn(value_label, format="$%.2f", disabled=True),
                    "Payout": st.column_config.NumberColumn("Payout", format="$%.0f", disabled=True),
                    "P&L": st.column_config.NumberColumn("P&L", format="$%.0f", disabled=True),
                },
                hide_index=True,
                use_container_width=True,
                key="puts_editor"
            )

            # Save positions and detect changes
            puts_changed = False
            for idx, row in edited_puts.iterrows():
                strike = row['Strike']
                old_pos = st.session_state['puts_positions'].get(strike, {})
                if row['Position'] != 0:
                    new_pos = {'position': row['Position'], 'entry': row['Entry']}
                    if old_pos != new_pos:
                        puts_changed = True
                    st.session_state['puts_positions'][strike] = new_pos
                elif strike in st.session_state['puts_positions']:
                    puts_changed = True
                    del st.session_state['puts_positions'][strike]

            if puts_changed:
                st.rerun()

        # Download button
        st.divider()
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            excel_file = create_excel_download(
                ticker, calls_pnl, puts_pnl, expiry, current_price,
                st.session_state['expirations']
            )
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            st.download_button(
                label="📥 Download Excel",
                data=excel_file,
                file_name=f"{ticker}_OPTIONS_{expiry}_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    else:
        st.error("Failed to load options chain")

else:
    st.info("👈 Enter a ticker symbol and click **Load Expirations** to get started")

    st.markdown("""
    ### How to use:
    1. Enter a stock ticker (e.g., CCJ, AAPL, TSLA)
    2. Click **Load Expirations** to see available option dates
    3. Select an expiration from the dropdown
    4. **Enter positions** directly in the table (+ for long, - for short)
    5. Use **Scenario Analysis** in sidebar:
       - **Days to Scenario**: Set to 0 for expiration P&L, or >0 for pre-expiry modeling
       - **Stock Price @ Scenario**: Expected stock price at scenario date
       - **IV Adjustment**: Model changes in implied volatility
    6. Optionally add a **Stock Position** (shares)
    7. View real-time **P&L Summary** at the top

    ### Scenario Analysis:
    - **At Expiry (Days=0)**: Shows intrinsic value (actual payoff at expiration)
    - **Before Expiry (Days>0)**: Uses Black-Scholes to model theoretical option values
    - **IV Adjustment**: Increase/decrease IV to see impact on option prices

    ### P&L Breakdown:
    - **Premiums Paid**: Cost of long positions
    - **Premiums Received**: Income from short positions
    - **Options Payout**: Value (intrinsic or theoretical) × position
    - **Stock P&L**: Profit/loss on stock position
    - **Total P&L**: Combined result
    """)
