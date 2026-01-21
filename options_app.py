#!/usr/bin/env python3
"""
Options Pricing App with GUI
Select ticker and expiration from dropdowns, generates Excel file.
"""

import tkinter as tk
from tkinter import ttk, messagebox
import threading
import math
from datetime import datetime
from pathlib import Path
import subprocess
import sys

try:
    import yfinance as yf
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing required package. Run: pip3 install yfinance pandas openpyxl")
    sys.exit(1)

# Try scipy, fall back to math.erf
try:
    from scipy.stats import norm
    def normal_cdf(x):
        return norm.cdf(x)
except ImportError:
    def normal_cdf(x):
        return 0.5 * (1 + math.erf(x / math.sqrt(2)))

# Constants
RISK_FREE_RATE = 0.045
DARK_HEADER_FILL = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")
BLUE_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
YELLOW_INPUT_FILL = PatternFill(start_color="FFFFC8", end_color="FFFFC8", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def calculate_delta(spot, strike, time_to_expiry, volatility, option_type="CALL"):
    if time_to_expiry <= 0 or volatility <= 0 or spot <= 0 or strike <= 0:
        return 0.5 if option_type == "CALL" else -0.5
    try:
        d1 = (math.log(spot / strike) + (RISK_FREE_RATE + 0.5 * volatility ** 2) * time_to_expiry) / (volatility * math.sqrt(time_to_expiry))
        return normal_cdf(d1) if option_type == "CALL" else normal_cdf(d1) - 1
    except:
        return 0.5 if option_type == "CALL" else -0.5


def create_options_sheet(ws, options_df, option_type, expiry, current_price, ticker):
    expiry_date = datetime.strptime(expiry, "%Y-%m-%d")
    days_to_expiry = (expiry_date - datetime.now()).days
    time_to_expiry = max(days_to_expiry / 365.0, 0.001)

    # === ROW 1: Market Data Headers ===
    for col, text in [(2, "Ticker"), (3, "Expiry"), (4, "Days"), (6, "Current Price"), (8, "Stock @ Expiry")]:
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = DARK_HEADER_FILL if col <= 6 else BLUE_HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER

    # === ROW 2: Market Data Values ===
    ws.cell(row=2, column=2, value=ticker).border = THIN_BORDER
    ws.cell(row=2, column=3, value=expiry).border = THIN_BORDER
    ws.cell(row=2, column=4, value=days_to_expiry).border = THIN_BORDER
    ws.cell(row=2, column=6, value=current_price).number_format = '$#,##0.00'
    ws.cell(row=2, column=6).border = THIN_BORDER

    exp_cell = ws.cell(row=2, column=8, value=current_price)
    exp_cell.number_format = '$#,##0.00'
    exp_cell.fill = YELLOW_INPUT_FILL
    exp_cell.border = THIN_BORDER

    # === ROW 1-2: Stock Position Input (columns J-L) ===
    ws.cell(row=1, column=10, value="Stock Shares").fill = BLUE_HEADER_FILL
    ws.cell(row=1, column=10).font = WHITE_FONT
    ws.cell(row=1, column=10).border = THIN_BORDER

    ws.cell(row=1, column=11, value="Stock Entry").fill = BLUE_HEADER_FILL
    ws.cell(row=1, column=11).font = WHITE_FONT
    ws.cell(row=1, column=11).border = THIN_BORDER

    ws.cell(row=1, column=12, value="Stock P&L").fill = BLUE_HEADER_FILL
    ws.cell(row=1, column=12).font = WHITE_FONT
    ws.cell(row=1, column=12).border = THIN_BORDER

    # Stock shares input (+ = long, - = short)
    stock_shares_cell = ws.cell(row=2, column=10, value=0)
    stock_shares_cell.fill = YELLOW_INPUT_FILL
    stock_shares_cell.border = THIN_BORDER

    # Stock entry price input
    stock_entry_cell = ws.cell(row=2, column=11, value=current_price)
    stock_entry_cell.number_format = '$#,##0.00'
    stock_entry_cell.fill = YELLOW_INPUT_FILL
    stock_entry_cell.border = THIN_BORDER

    # Stock P&L formula: (Stock@Expiry - StockEntry) * Shares
    ws.cell(row=2, column=12, value="=(H2-K2)*J2").number_format = '$#,##0.00'
    ws.cell(row=2, column=12).border = THIN_BORDER
    ws.cell(row=2, column=12).font = Font(bold=True)

    # === ROW 3: P&L Summary Section ===
    ws.cell(row=3, column=1, value="SUMMARY").font = Font(bold=True, size=11)

    # Labels and value cells side by side
    ws.cell(row=3, column=2, value="Premiums Paid:").font = Font(bold=True)
    # Value in col 3 (formula added after data)

    ws.cell(row=3, column=4, value="Premiums Rcvd:").font = Font(bold=True)
    # Value in col 5

    ws.cell(row=3, column=6, value="Options Payout:").font = Font(bold=True)
    # Value in col 7

    ws.cell(row=3, column=8, value="Options P&L:").font = Font(bold=True)
    # Value in col 9

    ws.cell(row=3, column=10, value="Stock P&L:").font = Font(bold=True)
    ws.cell(row=3, column=11, value="=L2").number_format = '$#,##0.00'  # Reference stock P&L
    ws.cell(row=3, column=11).font = Font(bold=True)

    ws.cell(row=3, column=12, value="TOTAL P&L:").font = Font(bold=True, color="0070C0")
    # Value in col 13 (Options P&L + Stock P&L)

    # === ROW 5: Column Headers (data starts row 6) ===
    headers = [
        ("Strike", DARK_HEADER_FILL),
        ("Bid", DARK_HEADER_FILL),
        ("Ask", DARK_HEADER_FILL),
        ("Last", DARK_HEADER_FILL),
        ("Mid", DARK_HEADER_FILL),
        ("Volume", DARK_HEADER_FILL),
        ("Open Int", DARK_HEADER_FILL),
        ("Impl Vol", DARK_HEADER_FILL),
        ("Delta", DARK_HEADER_FILL),
        ("Position", BLUE_HEADER_FILL),
        ("Entry", BLUE_HEADER_FILL),
        ("Prem Paid", BLUE_HEADER_FILL),
        ("Prem Rcvd", BLUE_HEADER_FILL),
        ("Val@Exp", BLUE_HEADER_FILL),
        ("Payout", BLUE_HEADER_FILL),
        ("P&L", BLUE_HEADER_FILL),
    ]
    for col, (h, fill) in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER

    row = 6
    for _, r in options_df.iterrows():
        strike = r.get('strike', 0)
        bid = r.get('bid', 0) or 0
        ask = r.get('ask', 0) or 0
        last = r.get('lastPrice', 0) or 0
        mid = (bid + ask) / 2 if bid > 0 and ask > 0 else last
        vol = r.get('volume', 0) or 0
        oi = r.get('openInterest', 0) or 0
        iv = r.get('impliedVolatility', 0) or 0

        # Market data (A-I)
        ws.cell(row=row, column=1, value=strike).number_format = '$#,##0.00'
        ws.cell(row=row, column=2, value=bid).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=ask).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=last).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=mid).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=int(vol) if pd.notna(vol) else 0).number_format = '#,##0'
        ws.cell(row=row, column=7, value=int(oi) if pd.notna(oi) else 0).number_format = '#,##0'
        ws.cell(row=row, column=8, value=iv if pd.notna(iv) else 0).number_format = '0.00%'

        delta = calculate_delta(current_price, strike, time_to_expiry, iv, option_type) if pd.notna(iv) and iv > 0 else 0
        ws.cell(row=row, column=9, value=delta).number_format = '0.00'

        # Position (J) - user input
        pos_cell = ws.cell(row=row, column=10, value=0)
        pos_cell.fill = YELLOW_INPUT_FILL

        # Entry Price (K) - user input
        entry_cell = ws.cell(row=row, column=11, value=mid)
        entry_cell.number_format = '$#,##0.00'
        entry_cell.fill = YELLOW_INPUT_FILL

        # Premium Paid (L) = Entry * Position * 100 if Position > 0, else 0
        ws.cell(row=row, column=12, value=f"=IF(J{row}>0,K{row}*J{row}*100,0)").number_format = '$#,##0.00'

        # Premium Received (M) = Entry * -Position * 100 if Position < 0, else 0
        ws.cell(row=row, column=13, value=f"=IF(J{row}<0,K{row}*-J{row}*100,0)").number_format = '$#,##0.00'

        # Value at Expiry per contract (N)
        if option_type == "CALL":
            ws.cell(row=row, column=14, value=f"=MAX($H$2-A{row},0)").number_format = '$#,##0.00'
        else:
            ws.cell(row=row, column=14, value=f"=MAX(A{row}-$H$2,0)").number_format = '$#,##0.00'

        # Payout (O) = Val@Exp * Position * 100
        ws.cell(row=row, column=15, value=f"=N{row}*J{row}*100").number_format = '$#,##0.00'

        # P&L (P) = Payout - Premium Paid + Premium Received
        ws.cell(row=row, column=16, value=f"=O{row}-L{row}+M{row}").number_format = '$#,##0.00'

        row += 1

    last_row = row - 1

    # === Summary formulas (Row 3) ===
    # Total Premiums Paid
    ws.cell(row=3, column=3, value=f"=SUM(L6:L{last_row})").number_format = '$#,##0.00'
    ws.cell(row=3, column=3).font = Font(bold=True)

    # Total Premiums Received
    ws.cell(row=3, column=5, value=f"=SUM(M6:M{last_row})").number_format = '$#,##0.00'
    ws.cell(row=3, column=5).font = Font(bold=True)

    # Total Options Payout
    ws.cell(row=3, column=7, value=f"=SUM(O6:O{last_row})").number_format = '$#,##0.00'
    ws.cell(row=3, column=7).font = Font(bold=True)

    # Options P&L
    ws.cell(row=3, column=9, value=f"=SUM(P6:P{last_row})").number_format = '$#,##0.00'
    ws.cell(row=3, column=9).font = Font(bold=True)

    # Total P&L (Options + Stock)
    total_pnl = ws.cell(row=3, column=13, value="=I3+K3")
    total_pnl.number_format = '$#,##0.00'
    total_pnl.font = Font(bold=True, size=12, color="0070C0")

    # Conditional formatting for P&L
    green_font = Font(color="008000")
    red_font = Font(color="C00000")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Row-level P&L
    ws.conditional_formatting.add(f"P6:P{last_row}", CellIsRule(operator='greaterThan', formula=['0'], font=green_font))
    ws.conditional_formatting.add(f"P6:P{last_row}", CellIsRule(operator='lessThan', formula=['0'], font=red_font))

    # Summary P&L cells
    for cell_ref in ["I3", "K3", "L2"]:
        ws.conditional_formatting.add(cell_ref, CellIsRule(operator='greaterThan', formula=['0'], font=green_font))
        ws.conditional_formatting.add(cell_ref, CellIsRule(operator='lessThan', formula=['0'], font=red_font))

    # Total P&L with background highlight
    ws.conditional_formatting.add("M3", CellIsRule(operator='greaterThan', formula=['0'], font=green_font, fill=green_fill))
    ws.conditional_formatting.add("M3", CellIsRule(operator='lessThan', formula=['0'], font=red_font, fill=red_fill))

    # Column widths
    widths = [10, 8, 8, 8, 8, 9, 9, 9, 7, 8, 8, 10, 10, 9, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A6'


class OptionsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Options Pricing Tool")
        self.root.geometry("400x300")

        self.expirations = []
        self.current_price = 0
        self.stock = None

        # Ticker
        ttk.Label(root, text="Ticker:", font=('Arial', 12)).pack(pady=(20, 5))
        self.ticker_var = tk.StringVar(value="CCJ")
        ticker_frame = ttk.Frame(root)
        ticker_frame.pack()
        self.ticker_entry = ttk.Entry(ticker_frame, textvariable=self.ticker_var, width=10, font=('Arial', 14))
        self.ticker_entry.pack(side=tk.LEFT, padx=5)
        ttk.Button(ticker_frame, text="Load", command=self.load_expirations).pack(side=tk.LEFT)

        # Expiration dropdown
        ttk.Label(root, text="Expiration:", font=('Arial', 12)).pack(pady=(20, 5))
        self.expiry_var = tk.StringVar()
        self.expiry_combo = ttk.Combobox(root, textvariable=self.expiry_var, width=15, font=('Arial', 12), state='readonly')
        self.expiry_combo.pack()

        # Price display
        self.price_label = ttk.Label(root, text="", font=('Arial', 11))
        self.price_label.pack(pady=10)

        # Generate button
        self.generate_btn = ttk.Button(root, text="Generate Excel & Open", command=self.generate_excel)
        self.generate_btn.pack(pady=20)

        # Status
        self.status_var = tk.StringVar(value="Enter ticker and click Load")
        ttk.Label(root, textvariable=self.status_var, font=('Arial', 10)).pack(pady=10)

    def load_expirations(self):
        ticker = self.ticker_var.get().strip().upper()
        if not ticker:
            messagebox.showerror("Error", "Please enter a ticker")
            return

        self.status_var.set(f"Loading {ticker}...")
        self.root.update()

        def fetch():
            try:
                self.stock = yf.Ticker(ticker)
                self.expirations = list(self.stock.options)

                try:
                    self.current_price = self.stock.info.get('regularMarketPrice') or self.stock.info.get('currentPrice') or 0
                    if not self.current_price:
                        hist = self.stock.history(period="1d")
                        self.current_price = hist['Close'].iloc[-1] if not hist.empty else 0
                except:
                    self.current_price = 0

                self.root.after(0, self.update_ui)
            except Exception as e:
                self.root.after(0, lambda: self.status_var.set(f"Error: {e}"))

        threading.Thread(target=fetch, daemon=True).start()

    def update_ui(self):
        if self.expirations:
            self.expiry_combo['values'] = self.expirations
            self.expiry_combo.current(0)
            self.price_label.config(text=f"Current Price: ${self.current_price:.2f}")
            self.status_var.set(f"Found {len(self.expirations)} expirations")
        else:
            self.status_var.set("No options found for this ticker")

    def generate_excel(self):
        ticker = self.ticker_var.get().strip().upper()
        expiry = self.expiry_var.get()

        if not ticker or not expiry:
            messagebox.showerror("Error", "Please load a ticker and select expiration")
            return

        self.status_var.set("Generating Excel...")
        self.root.update()

        def generate():
            try:
                chain = self.stock.option_chain(expiry)
                calls_df = chain.calls
                puts_df = chain.puts

                wb = Workbook()
                ws_calls = wb.active
                ws_calls.title = "Calls"
                create_options_sheet(ws_calls, calls_df, "CALL", expiry, self.current_price, ticker)

                ws_puts = wb.create_sheet("Puts")
                create_options_sheet(ws_puts, puts_df, "PUT", expiry, self.current_price, ticker)

                # Expirations tab
                ws_exp = wb.create_sheet("Expirations")
                ws_exp.cell(row=1, column=1, value="Available Expirations").font = Font(bold=True)
                for i, exp in enumerate(self.expirations, 3):
                    cell = ws_exp.cell(row=i, column=1, value=exp)
                    if exp == expiry:
                        cell.font = Font(bold=True, color="0070C0")

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{ticker}_OPTIONS_{expiry}_{timestamp}.xlsx"
                filepath = Path.cwd() / filename
                wb.save(str(filepath))

                # Open the file
                subprocess.run(['open', str(filepath)])

                self.root.after(0, lambda: self.status_var.set(f"Created: {filename}"))

            except Exception as e:
                self.root.after(0, lambda: self.status_var.set(f"Error: {e}"))

        threading.Thread(target=generate, daemon=True).start()


def main():
    root = tk.Tk()
    app = OptionsApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
